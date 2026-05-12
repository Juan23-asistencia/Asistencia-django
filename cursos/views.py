import datetime, calendar as cal_mod
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Curso, Materia, Estudiante
from .forms import CursoForm, MateriaForm, EstudianteForm


# ─── CURSOS ──────────────────────────────────────────────────────────────────

@login_required
def lista_cursos(request):
    """
    Admin: todos los cursos agrupados por nivel.
    Profesor: solo los cursos donde tiene al menos una materia asignada.
    """
    user = request.user
    if user.es_admin():
        cursos = Curso.objects.prefetch_related('materias', 'estudiantes').all()
    else:
        # cursos donde el profesor tiene alguna materia
        ids = Materia.objects.filter(profesor=user).values_list('curso_id', flat=True).distinct()
        cursos = Curso.objects.filter(pk__in=ids).prefetch_related('materias', 'estudiantes')

    # Agrupar por nivel en orden fijo
    ORDEN = ['10mo', '1ro_bach', '2do_bach', '3ro_bach']
    grupos = {}
    for c in cursos:
        key = c.nivel
        grupos.setdefault(key, []).append(c)

    # Ordenar grupos según ORDEN y convertir key a label para el template
    grupos_ordenados = {}
    for key in ORDEN:
        if key in grupos:
            label = grupos[key][0].get_nivel_display()
            grupos_ordenados[label] = sorted(grupos[key], key=lambda x: x.paralelo)

    return render(request, 'cursos/lista_cursos.html', {'grupos': grupos_ordenados, 'cursos': cursos})


@login_required
def detalle_curso(request, pk):
    """
    Muestra las materias del curso.
    Admin: ve todas.  Profesor: solo las suyas.
    """
    user = request.user
    curso = get_object_or_404(Curso, pk=pk)

    if user.es_admin():
        materias = curso.materias.select_related('profesor').all()
    else:
        materias = curso.materias.filter(profesor=user).select_related('profesor')
        if not materias.exists():
            messages.error(request, 'No tienes materias asignadas en este curso.')
            return redirect('lista_cursos')

    estudiantes = curso.estudiantes.all()
    import json
    materias_json = json.dumps([m.nombre for m in materias])
    return render(request, 'cursos/detalle_curso.html', {
        'curso':        curso,
        'materias':     materias,
        'estudiantes':  estudiantes,
        'materias_json': materias_json,
    })


@login_required
def crear_curso(request):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            curso = form.save()
            messages.success(request, f'Curso {curso} creado.')
            return redirect('detalle_curso', pk=curso.pk)
    else:
        form = CursoForm()
    return render(request, 'cursos/form_curso.html', {'form': form, 'accion': 'Crear'})


@login_required
def editar_curso(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso actualizado.')
            return redirect('lista_cursos')
    else:
        form = CursoForm(instance=curso)
    return render(request, 'cursos/form_curso.html', {'form': form, 'accion': 'Editar', 'curso': curso})


@login_required
def eliminar_curso(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        curso.delete()
        messages.success(request, f'Curso eliminado.')
        return redirect('lista_cursos')
    return render(request, 'cursos/confirmar_eliminar_curso.html', {'curso': curso})


# ─── MATERIAS ────────────────────────────────────────────────────────────────

@login_required
def detalle_materia(request, pk):
    """
    Vista principal de la materia: lista de estudiantes + grilla de asistencia.
    Solo puede entrar el profesor asignado o el admin.
    """
    from asistencia.models import Asistencia

    user   = request.user
    materia = get_object_or_404(Materia, pk=pk)

    if not user.es_admin() and materia.profesor != user:
        messages.error(request, 'No tienes acceso a esta materia.')
        return redirect('lista_cursos')

    curso       = materia.curso
    hoy         = datetime.date.today()
    mes         = int(request.GET.get('mes',  hoy.month))
    anio        = int(request.GET.get('anio', hoy.year))
    num_dias    = cal_mod.monthrange(anio, mes)[1]
    fechas_mes  = [datetime.date(anio, mes, d) for d in range(1, num_dias + 1)]

    # Asistencias del mes para esta materia
    asist_qs = Asistencia.objects.filter(
        materia=materia, fecha__year=anio, fecha__month=mes
    ).values('estudiante_id', 'fecha__day', 'estado')

    mapa = {}
    for a in asist_qs:
        mapa.setdefault(a['estudiante_id'], {})[a['fecha__day']] = a['estado']

    estudiantes = curso.estudiantes.all()
    filas = []
    for est in estudiantes:
        dias_row  = []
        presentes = ausentes = 0
        for fecha in fechas_mes:
            estado = mapa.get(est.pk, {}).get(fecha.day)
            finde  = fecha.weekday() >= 5
            if estado == 'presente': presentes += 1
            elif estado == 'ausente': ausentes += 1
            dias_row.append({'dia': fecha.day, 'estado': estado, 'finde': finde})
        total = presentes + ausentes
        filas.append({
            'estudiante': est,
            'dias':       dias_row,
            'presentes':  presentes,
            'ausentes':   ausentes,
            'pct':        round(presentes / total * 100) if total else None,
        })

    meses_nombres = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                     'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    meses_lista = [(i, meses_nombres[i]) for i in range(1, 13)]
    anios_lista = list(range(hoy.year - 2, hoy.year + 1))

    mes_ant  = 12 if mes == 1  else mes - 1
    anio_ant = anio - 1 if mes == 1  else anio
    mes_sig  = 1  if mes == 12 else mes + 1
    anio_sig = anio + 1 if mes == 12 else anio

    return render(request, 'cursos/detalle_materia.html', {
        'materia':     materia,
        'curso':       curso,
        'filas':       filas,
        'fechas_mes':  fechas_mes,
        'mes': mes, 'anio': anio,
        'nombre_mes':  meses_nombres[mes],
        'meses_lista': meses_lista,
        'anios_lista': anios_lista,
        'mes_ant': mes_ant, 'anio_ant': anio_ant,
        'mes_sig': mes_sig, 'anio_sig': anio_sig,
        'hoy': hoy,
    })


@login_required
def crear_materia(request, curso_pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    curso = get_object_or_404(Curso, pk=curso_pk)
    if request.method == 'POST':
        form = MateriaForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.curso = curso
            m.save()
            messages.success(request, f'Materia «{m.nombre}» creada.')
            return redirect('detalle_curso', pk=curso_pk)
    else:
        form = MateriaForm()
    return render(request, 'cursos/form_materia.html', {'form': form, 'curso': curso, 'accion': 'Crear'})


@login_required
def editar_materia(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    materia = get_object_or_404(Materia, pk=pk)
    if request.method == 'POST':
        form = MateriaForm(request.POST, instance=materia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Materia actualizada.')
            return redirect('detalle_curso', pk=materia.curso_id)
    else:
        form = MateriaForm(instance=materia)
    return render(request, 'cursos/form_materia.html', {'form': form, 'curso': materia.curso, 'accion': 'Editar', 'materia': materia})


@login_required
def eliminar_materia(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    materia = get_object_or_404(Materia, pk=pk)
    curso_pk = materia.curso_id
    if request.method == 'POST':
        materia.delete()
        messages.success(request, 'Materia eliminada.')
        return redirect('detalle_curso', pk=curso_pk)
    return render(request, 'cursos/confirmar_eliminar_materia.html', {'materia': materia})


# ─── ESTUDIANTES ─────────────────────────────────────────────────────────────

@login_required
def agregar_estudiante(request, curso_pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    curso = get_object_or_404(Curso, pk=curso_pk)
    if request.method == 'POST':
        form = EstudianteForm(request.POST)
        if form.is_valid():
            est = form.save(commit=False)
            est.curso = curso
            est.save()
            messages.success(request, f'Estudiante {est} agregado.')
            return redirect('detalle_curso', pk=curso_pk)
    else:
        form = EstudianteForm()
    return render(request, 'cursos/form_estudiante.html', {'form': form, 'curso': curso, 'accion': 'Agregar'})


@login_required
def editar_estudiante(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    estudiante = get_object_or_404(Estudiante, pk=pk)
    if request.method == 'POST':
        form = EstudianteForm(request.POST, instance=estudiante)
        if form.is_valid():
            form.save()
            messages.success(request, 'Estudiante actualizado.')
            return redirect('detalle_curso', pk=estudiante.curso_id)
    else:
        form = EstudianteForm(instance=estudiante)
    return render(request, 'cursos/form_estudiante.html', {'form': form, 'curso': estudiante.curso, 'accion': 'Editar'})


@login_required
def eliminar_estudiante(request, pk):
    if not request.user.es_admin():
        messages.error(request, 'Sin permiso.'); return redirect('dashboard')
    estudiante = get_object_or_404(Estudiante, pk=pk)
    curso_pk   = estudiante.curso_id
    if request.method == 'POST':
        estudiante.delete()
        messages.success(request, 'Estudiante eliminado.')
        return redirect('detalle_curso', pk=curso_pk)
    return render(request, 'cursos/confirmar_eliminar_estudiante.html', {'estudiante': estudiante})


# ─── AJAX: agregar estudiante desde detalle_curso ────────────────────────────

@login_required
def agregar_estudiante_ajax(request, curso_pk):
    import json
    if not request.user.es_admin():
        return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

    curso = get_object_or_404(Curso, pk=curso_pk)

    if request.method == 'POST':
        nombre   = request.POST.get('nombre',   '').strip()
        apellido = request.POST.get('apellido', '').strip()
        cedula   = request.POST.get('cedula',   '').strip() or None

        if not nombre or not apellido:
            return JsonResponse({'ok': False, 'error': 'Nombre y apellido son obligatorios'})

        if cedula and Estudiante.objects.filter(cedula=cedula).exists():
            return JsonResponse({'ok': False, 'error': f'Ya existe un estudiante con cédula {cedula}'})

        est = Estudiante.objects.create(
            nombre=nombre, apellido=apellido,
            cedula=cedula, curso=curso
        )
        return JsonResponse({
            'ok':      True,
            'id':      est.pk,
            'nombre':  est.nombre,
            'apellido': est.apellido,
            'cedula':  est.cedula or '',
        })

    return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)



# ─── IMPORTAR ESTUDIANTES DESDE EXCEL ────────────────────────────────────────

@login_required
def importar_estudiantes(request, curso_pk):
    """Recibe un archivo Excel/CSV y crea los estudiantes en el curso."""
    if not request.user.es_admin():
        return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

    curso = get_object_or_404(Curso, pk=curso_pk)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'ok': False, 'error': 'No se recibió ningún archivo'})

    nombre = archivo.name.lower()
    if not (nombre.endswith('.xlsx') or nombre.endswith('.xls') or nombre.endswith('.csv')):
        return JsonResponse({'ok': False, 'error': 'Solo se aceptan .xlsx, .xls o .csv'})

    try:
        import io, csv
        filas = []

        if nombre.endswith('.csv'):
            contenido = archivo.read().decode('utf-8-sig', errors='replace')
            reader = csv.reader(io.StringIO(contenido))
            todas = list(reader)
            if not todas:
                return JsonResponse({'ok': False, 'error': 'El archivo CSV está vacío'})
            # Primera fila: encabezados
            enc = [str(h).strip().lower() for h in todas[0]]
            for fila in todas[1:]:
                if not any(c.strip() for c in fila):
                    continue
                filas.append(dict(zip(enc, [c.strip() for c in fila])))
        else:
            import openpyxl
            wb   = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return JsonResponse({'ok': False, 'error': 'El archivo Excel está vacío'})
            enc = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
            for row in rows[1:]:
                if all(c is None for c in row):
                    continue
                vals = [str(c).strip() if c is not None else '' for c in row]
                filas.append(dict(zip(enc, vals)))

        if not filas:
            return JsonResponse({'ok': False, 'error': 'No se encontraron datos en el archivo'})

        # ── Detectar columnas de forma flexible ──────────────────────────────
        ALIAS = {
            'apellido': ['apellido', 'apellidos', 'lastname', 'last_name',
                         'surname', 'primer apellido'],
            'nombre':   ['nombre', 'nombres', 'name', 'names',
                         'primer nombre', 'firstname', 'first_name'],
            'cedula':   ['cedula', 'cédula', 'ci', 'identificacion',
                         'identificación', 'id', 'documento'],
        }

        def detectar_col(fila_keys, campo):
            for k in fila_keys:
                if k in ALIAS[campo]:
                    return k
            return None

        keys_muestra  = list(filas[0].keys()) if filas else []
        col_apellido  = detectar_col(keys_muestra, 'apellido')
        col_nombre    = detectar_col(keys_muestra, 'nombre')
        col_cedula    = detectar_col(keys_muestra, 'cedula')

        agregados  = []
        duplicados = []
        errores    = []

        for i, fila in enumerate(filas, start=2):
            # Obtener valores por columna detectada o por posición
            vals = list(fila.values())

            if col_apellido and col_nombre:
                apellido = fila.get(col_apellido, '').strip()
                nombre   = fila.get(col_nombre,   '').strip()
                cedula   = fila.get(col_cedula,   '').strip() if col_cedula else ''
            else:
                # Sin encabezados reconocidos: usar posición 0=apellido, 1=nombre, 2=cedula
                apellido = vals[0].strip() if len(vals) > 0 else ''
                nombre   = vals[1].strip() if len(vals) > 1 else ''
                cedula   = vals[2].strip() if len(vals) > 2 else ''

            # Limpiar cédula
            cedula = cedula if cedula and cedula.lower() not in ('none', 'n/a', '-', '') else None

            if not apellido or not nombre:
                if any(v.strip() for v in vals if v):
                    errores.append(f'Fila {i}: falta nombre o apellido')
                continue

            # Verificar duplicados
            if cedula and Estudiante.objects.filter(cedula=cedula).exclude(
                    nombre__iexact=nombre, apellido__iexact=apellido, curso=curso).exists():
                duplicados.append(f'{apellido}, {nombre} — cédula {cedula} ya existe')
                continue

            if Estudiante.objects.filter(
                    nombre__iexact=nombre,
                    apellido__iexact=apellido,
                    curso=curso).exists():
                duplicados.append(f'{apellido}, {nombre} — ya está en el curso')
                continue

            est = Estudiante.objects.create(
                nombre=nombre, apellido=apellido,
                cedula=cedula, curso=curso
            )
            agregados.append({
                'id':      est.pk,
                'nombre':  est.nombre,
                'apellido': est.apellido,
                'cedula':  est.cedula or '',
            })

        return JsonResponse({
            'ok':         True,
            'agregados':  agregados,
            'duplicados': duplicados,
            'errores':    errores,
            'total':      len(agregados),
        })

    except Exception as e:
        import traceback
        return JsonResponse({
            'ok':    False,
            'error': f'Error procesando el archivo: {str(e)}',
        })


# ─── DESCARGAR PLANTILLA EXCEL ───────────────────────────────────────────────

@login_required
def descargar_plantilla_excel(request):
    try:
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"

        # Encabezados con estilo
        from openpyxl.styles import Font, PatternFill, Alignment
        headers = ['apellido', 'nombre', 'cedula']
        header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFF', size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos de ejemplo
        ejemplos = [
            ('Pérez',    'Juan',     '0912345671'),
            ('García',   'María',    '0923456782'),
            ('Torres',   'Carlos',   '0934567893'),
            ('Rodríguez','Ana',      '0945678904'),
            ('Mendoza',  'Lucía',    ''),
        ]
        for fila, (ap, nom, ced) in enumerate(ejemplos, 2):
            ws.cell(row=fila, column=1, value=ap)
            ws.cell(row=fila, column=2, value=nom)
            ws.cell(row=fila, column=3, value=ced)

        # Ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.xlsx"'
        wb.save(response)
        return response

    except ImportError:
        from django.http import HttpResponse
        # Si no está openpyxl, devolver CSV
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="plantilla_estudiantes.csv"'
        response.write('\ufeff')  # BOM para Excel
        writer = csv.writer(response)
        writer.writerow(['apellido', 'nombre', 'cedula'])
        writer.writerow(['Pérez', 'Juan', '0912345671'])
        writer.writerow(['García', 'María', '0923456782'])
        writer.writerow(['Torres', 'Carlos', ''])
        return response
